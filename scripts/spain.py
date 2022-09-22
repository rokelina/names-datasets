import os
from openpyxl import load_workbook
from name_classes import SpainFemaleNames, SpainMaleNames
from typing import List


def get_list_of_workbooks(directory_path):
    list_of_workbooks = []
    for file in os.listdir(directory_path):
        if file.endswith('.xlsx'):
            wb = load_workbook(os.path.join(directory_path, file))
            list_of_workbooks.append(wb)
    return list_of_workbooks


def write_dataset_year(workbooks):
    for wb in workbooks:
        year = wb.active.title
        sheet = wb['TOTAL']
        for row in sheet['C5:C105']:
            dataset_year = int((year.split(' ')[1]))
            for cell in row:
                cell.value = dataset_year
    return workbooks


def get_main_worksheets(workbooks):
    list_of_worksheets = []
    for wb in workbooks:
        sheet = wb['TOTAL']
        list_of_worksheets.append(sheet)
    return list_of_worksheets


def get_rows_of_names(worksheets):
    list_of_names_data = []
    for ws in worksheets:
        for row in ws.iter_rows(min_row=5, max_row=104, min_col=1, max_col=5, values_only=True):
            # iter_rows returns each row as a tuple. Transform it into a list before appending
            list_of_names_data.append(list(row))
    return list_of_names_data


spain_dir = "/Users/rosinascampino/Desktop/names_project/raw_data/spain"
all_wbs = write_dataset_year(get_list_of_workbooks(spain_dir))
spain_data = get_rows_of_names(get_main_worksheets(all_wbs))


spain_MaleNames_objects: List[SpainMaleNames] = []
spain_FemaleNames_objects: List[SpainFemaleNames] = []

for row in spain_data:
    m_object = SpainMaleNames(row[0].title(), int(row[1]), row[2])
    f_object = SpainFemaleNames(row[3].title(), int(row[4]), row[2])
    spain_MaleNames_objects.append(m_object)
    spain_FemaleNames_objects.append(f_object)


all_names = [i.as_array() for i in spain_MaleNames_objects] + [i.as_array()
                                                               for i in spain_FemaleNames_objects]
