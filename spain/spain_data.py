from openpyxl import load_workbook
import csv
import os
from names_classes.spain_names import SpainFemaleNames, SpainMaleNames
from typing import List

spain_dir = "/Users/rosinascampino/Desktop/names_project/spain"


def get_list_of_workbooks(directory_path):
    '''Iterates over a directory and extracts all the worksheets that contain the total name count for that year'''

    list_of_workbooks = []
    for file in os.listdir(directory_path):
        if file.endswith('.xlsx'):
            wb = load_workbook(os.path.join(directory_path, file))
            list_of_workbooks.append(wb)
    return list_of_workbooks


# def write_year

# sheet = wb['TOTAL']

# '''Write the file year into each row'''
# for row in sheet['C5:C105']:
#     file_year = int((file.split('.')[0]))
#     for cell in row:
#         cell.value = file_year


def get_names_data(worksheets):
    '''Extracts all the rows of names from a list of worksheets'''

    list_of_names_data = []
    for ws in worksheets:
        for row in ws.iter_rows(min_row=5, max_row=104, min_col=1, max_col=5, values_only=True):
            list_of_names_data.append(list(row))
    return list_of_names_data


def write_csv(list_of_data):
    os.chdir('/Users/rosinascampino/Desktop/names_project/cleaned_data')
    with open('spain_names.csv', 'w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(['Name', 'Count', 'Year',
                        'Gender', 'Country'])
        writer.writerows(list_of_data)


spain_data = get_names_data(get_list_of_workbooks(spain_dir))

list_of_MaleNames_objects: List[SpainMaleNames] = []
list_of_FemaleNames_objects: List[SpainFemaleNames] = []

for row in spain_data:
    m_object = SpainMaleNames(row[0].title(), int(row[1]), row[2])
    f_object = SpainFemaleNames(row[3].title(), int(row[4]), row[2])
    list_of_MaleNames_objects.append(m_object)
    list_of_FemaleNames_objects.append(f_object)


all_names = [i.as_array() for i in list_of_MaleNames_objects] + [i.as_array()
                                                                 for i in list_of_FemaleNames_objects]

write_csv(all_names)
